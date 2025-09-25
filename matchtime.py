"""

The .trc files exported from QTM need to be manually converted to .xlsx files first. And then the code can run.

Time-matching script for your workflow.

- LOG_FILE: the Excel event log (SubjectID, Task Name, Trial, Timestamp, Event, ...)
- QTM_FOLDER: folder containing the converted QTM .xlsx files (filenames like Task_Trial1_SubjectID.xlsx)
- OUTPUT_FOLDER: where the *_synced.xlsx files will be saved
"""

import os
import re
import math
from datetime import timedelta
import pandas as pd
from openpyxl import load_workbook

# ---------------- CONFIG ----------------
LOG_FILE = r"C:\Users\AoMV Lab\ricki projects\trial_log_20250925_123040_rickitesting.xlsx"
QTM_FOLDER = r"C:\Users\Ricki\Documents\beep\beeptest\Data"
OUTPUT_FOLDER = r"C:\Users\AoMV Lab\ricki projects\TimeMatched"

# Which events to include from the log
EVENT_STATIC = [
    "QTM Start Command Sent",
    "QTM Recording Started",
    "Beep Started",
]
EVENT_PATTERNS = [
    r"LED_\d+_Lit",     # LED_1_Lit etc
    r"#\d+\s*-\s*pressed"  # #1 - pressed
]
# Maximum allowed match distance (seconds) before warning (set None to skip check)
MAX_MATCH_SECONDS = 0.5

# ----------------------------------------

def matches_event_name(s: str) -> bool:
    if s is None:
        return False
    s = str(s)
    if s.strip() in EVENT_STATIC:
        return True
    for pat in EVENT_PATTERNS:
        if re.search(pat, s, flags=re.IGNORECASE):
            return True
    return False


def find_header_row(ws, max_search_row=12):
    """Try to find the header row by looking for both 'frame' and 'time' in the same row.
       If not found, default to row 4 (your stated layout)."""
    for r in range(1, min(max_search_row, ws.max_row) + 1):
        vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            vals.append(str(v).lower() if v is not None else "")
        row_text = " ".join(vals)
        if "frame" in row_text and "time" in row_text:
            return r
    return 4


def find_col_by_keyword(ws, header_row, keywords):
    """Return first column index where header contains any of the keywords"""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        vl = str(v).lower()
        for kw in keywords:
            if kw in vl:
                return c
    return None


def parse_qtm_filename(fname):
    """Expect filenames like TaskName_Trial1_SubjectID.xlsx
       Returns (task, trial_number, subject) or None on fail."""
    base = os.path.splitext(os.path.basename(fname))[0]
    # main pattern:
    m = re.match(r'(?P<task>.+?)_Trial(?P<trial>\d+?)_(?P<subject>.+)$', base, flags=re.IGNORECASE)
    if m:
        return m.group('task'), int(m.group('trial')), m.group('subject')
    # fallback patterns (less strict)
    m2 = re.match(r'(?P<task>.+?)_trial[_\- ]?(?P<trial>\d+)[_\- ]?(?P<subject>.+)$', base, flags=re.IGNORECASE)
    if m2:
        return m2.group('task'), int(m2.group('trial')), m2.group('subject')
    return None


def safe_to_datetime(val):
    """Convert different timestamp formats to pandas.Timestamp"""
    if pd.isna(val):
        return None
    try:
        return pd.to_datetime(str(val))
    except Exception:
        # try forgiving parsing via pandas to_timedelta if it's a time-only like 16:09:17.422
        try:
            return pd.to_datetime(str(val), format="%H:%M:%S.%f")
        except Exception:
            return None


def parse_seconds_from_qtm_time(cell_val):
    """QTM Time column may be numeric seconds (0.000) or string like '00:00:00.008333'.
       Return seconds as float or None."""
    if cell_val is None:
        return None
    s = str(cell_val).strip()
    if s == "":
        return None
    # if looks like 'hh:mm:ss' or '00:00:00.008333', use to_timedelta
    if ":" in s:
        try:
            td = pd.to_timedelta(s)
            return td.total_seconds()
        except Exception:
            pass
    # otherwise numeric
    try:
        return float(s.replace(",", "."))
    except Exception:
        return None


def get_log_column_mapping(df):
    """Return mapping for subject/task/trial/timestamp/event headers as found in log_df."""
    mapping = {}
    for col in df.columns:
        lower = col.lower()
        if 'subject' in lower and 'id' in lower or ('subject' in lower and 'id' not in mapping):
            mapping['subject'] = col
        if 'task' in lower and 'name' in lower or ('task' in lower and 'task' not in mapping):
            mapping['task'] = col
        if 'trial' in lower and 'trial' not in mapping:
            mapping['trial'] = col
        if 'timestamp' in lower or ('time' == lower) or ('time' in lower and 'stamp' in lower) or ('time' in lower and 'timestamp' not in mapping):
            # pick a 'timestamp' like column
            mapping.setdefault('timestamp', col)
        if 'event' in lower and 'event' not in mapping:
            mapping['event'] = col
    # final checks
    keys = ['subject', 'task', 'trial', 'timestamp', 'event']
    for k in keys:
        if k not in mapping:
            raise KeyError(f"Could not find a '{k}' column in the log file. Columns available: {list(df.columns)}")
    return mapping


def process_single_qtm(qtm_path, trial_events_df, log_col_map):
    """Modify the qtm file in-place (save to OUTPUT_FOLDER with _synced)."""
    wb = load_workbook(qtm_path)
    ws = wb.active

    header_row = find_header_row(ws)
    data_start_row = header_row + 3  # you said data begins at row 7 when header is row 4

    # Insert Event column at A and Global Time Synced at D (row header_row)
    ws.insert_cols(1)  # new col A
    ws.insert_cols(4)  # new col D (after the first insert)

    # write header labels
    ws.cell(row=header_row, column=1).value = "Event"
    ws.cell(row=header_row, column=4).value = "Global Time Synced"

    # Find the Time column (after insertion we re-detect headers)
    time_col = find_col_by_keyword(ws, header_row, ['time'])
    if time_col is None:
        print(f"⚠ Could not find a 'Time' column in {os.path.basename(qtm_path)} (header row {header_row}). Skipping.")
        return

    # Obtain start_time from trial_events_df (find QTM Start Command Sent row)
    # trial_events_df column names are original names from log; map them using log_col_map
    ev_col = log_col_map['event']
    ts_col = log_col_map['timestamp']
    start_mask = trial_events_df[ev_col].astype(str).str.contains("QTM Start Command Sent", case=False, na=False)
    if not start_mask.any():
        print(f"⚠️ No 'QTM Start Command Sent' in log events for {os.path.basename(qtm_path)}. Skipping.")
        return
    start_time_raw = trial_events_df.loc[start_mask, ts_col].iloc[0]
    start_time = safe_to_datetime(start_time_raw)
    if start_time is None:
        print(f"⚠️ Could not parse start timestamp ({start_time_raw}) for {os.path.basename(qtm_path)}. Skipping.")
        return

    # Fill Global Time Synced for each data row and build a map row -> datetime
    row_to_global = {}
    for r in range(data_start_row, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=time_col).value
        secs = parse_seconds_from_qtm_time(cell_val)
        if secs is None or (isinstance(secs, float) and math.isnan(secs)):
            continue
        gtime = start_time + timedelta(seconds=secs)
        # write formatted time (HH:MM:SS.mmm)
        ws.cell(row=r, column=4).value = gtime.strftime("%H:%M:%S.%f")[:-3]
        row_to_global[r] = gtime.to_pydatetime() if hasattr(gtime, "to_pydatetime") else gtime

    if not row_to_global:
        print(f"⚠️ No time rows found in {os.path.basename(qtm_path)} (data start {data_start_row}). Skipping.")
        return

    # For each event in trial_events_df that matches our patterns, find nearest frame row
    events_of_interest_df = trial_events_df[trial_events_df[ev_col].apply(matches_event_name)]
    if events_of_interest_df.empty:
        print(f"⚠️ No matching event types found in log for {os.path.basename(qtm_path)}. Nothing to write.")
    else:
        # Parse all event timestamps and names
        ev_list = []
        for idx, row in events_of_interest_df.iterrows():
            ev_name = str(row[ev_col])
            ev_ts_raw = row[ts_col]
            ev_ts = safe_to_datetime(ev_ts_raw)
            if ev_ts is None:
                # try parsing with just time-of-day
                try:
                    ev_ts = pd.to_datetime(str(ev_ts_raw), format="%H:%M:%S.%f")
                except Exception:
                    print(f"⚠️ Could not parse event timestamp '{ev_ts_raw}' for event '{ev_name}'")
                    continue
            ev_list.append((ev_name, ev_ts.to_pydatetime() if hasattr(ev_ts, "to_pydatetime") else ev_ts))

        # match each event to the nearest row
        for ev_name, ev_dt in ev_list:
            # find nearest row by absolute time diff
            nearest_row = min(row_to_global.keys(), key=lambda r: abs((row_to_global[r] - ev_dt).total_seconds()))
            diff = abs((row_to_global[nearest_row] - ev_dt).total_seconds())
            if (MAX_MATCH_SECONDS is not None) and (diff > MAX_MATCH_SECONDS):
                print(f"⚠️ Large match difference for '{ev_name}' in {os.path.basename(qtm_path)}: {diff:.3f}s (row {nearest_row})")
            # append if already exists
            cur = ws.cell(row=nearest_row, column=1).value
            if cur and str(cur).strip():
                ws.cell(row=nearest_row, column=1).value = f"{cur} | {ev_name}"
            else:
                ws.cell(row=nearest_row, column=1).value = ev_name

    # Save output
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    out_name = os.path.splitext(os.path.basename(qtm_path))[0] + "_synced.xlsx"
    out_path = os.path.join(OUTPUT_FOLDER, out_name)
    wb.save(out_path)
    print(f"✅ Saved synced file: {out_path}")


def main():
    # read log
    print("Loading log:", LOG_FILE)
    log_df = pd.read_excel(LOG_FILE)
    log_map = get_log_column_mapping(log_df)

    # keep only rows that correspond to events we might care about (speed up)
    log_df_filtered = log_df[log_df[log_map['event']].apply(lambda s: matches_event_name(str(s))) | log_df[log_map['event']].astype(str).str.contains("QTM Start Command Sent", case=False, na=False)]

    # iterate qtm files
    for fname in sorted(os.listdir(QTM_FOLDER)):
        if not fname.lower().endswith(".xlsx"):
            continue
        if fname.startswith("~$"):  # skip Excel temp files
            continue
        if fname.lower().endswith("_synced.xlsx"):  # skip previously produced outputs
            continue
        qtm_path = os.path.join(QTM_FOLDER, fname)
        parsed = parse_qtm_filename(fname)
        if not parsed:
            print(f"⚠️ Could not parse filename '{fname}' using expected pattern. Skipping.")
            continue
        task, trial_num, subject = parsed
        # normalize
        task_norm = str(task).strip().lower()
        subject_norm = str(subject).strip().lower()

        # select rows from the log that match subject, task, and trial
        try:
            # convert types and compare robustly
            subj_col = log_map['subject']
            task_col = log_map['task']
            trial_col = log_map['trial']

            mask = (
                log_df_filtered[subj_col].astype(str).str.strip().str.lower() == subject_norm
            ) & (
                log_df_filtered[task_col].astype(str).str.strip().str.lower() == task_norm
            )
            # trial column numeric compare if possible
            try:
                mask = mask & (log_df_filtered[trial_col].astype(float).astype('Int64') == int(trial_num))
            except Exception:
                # fallback: string contains trial number
                mask = mask & (log_df_filtered[trial_col].astype(str).str.contains(str(trial_num)))
            trial_events = log_df_filtered[mask]
        except KeyError as e:
            print(f"⚠️ Problem mapping log columns for {fname}: {e}")
            continue

        if trial_events.empty:
            print(f"⚠️ No events found for {fname} (subject={subject}, task={task}, trial={trial_num})")
            continue

        process_single_qtm(qtm_path, trial_events, log_map)


if __name__ == "__main__":
    main()
