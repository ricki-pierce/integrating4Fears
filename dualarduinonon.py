#This code incorporates naming of tasks and trial number automatically in the terminal and in excel. this code does not automatically reset qtm to record a new trial.

import asyncio
import threading
import serial
import time
import random
import numpy as np
import sounddevice as sd
import tkinter as tk
from tkinter import simpledialog, messagebox, ttk
from openpyxl import Workbook
from datetime import datetime
import qtm
import os
import soundfile as sf
import subprocess
import pytz

# ------------------ CONFIG ------------------
SERIAL_PORT = 'COM10'
BAUD_RATE = 115200
WORK_DIR = r"C:\\Users\\AoMV Lab\\ricki projects"
os.chdir(WORK_DIR)

filename_beep = r"C:\\Users\\AoMV Lab\\ricki projects\\Silenceplus500hz1000mstone.wav"

# ------------------ TIME SYNC ------------------
def sync_windows_time():
    try:
        output = subprocess.check_output(['w32tm', '/resync'], shell=True, stderr=subprocess.STDOUT, text=True)
        print(f"Time sync success:\n{output}")
    except subprocess.CalledProcessError as e:
        print(f"Time sync failed:\n{e.output}")

CENTRAL_TZ = pytz.timezone('America/Chicago')

def now_central():
    return datetime.now(CENTRAL_TZ)

# ------------------ GLOBALS ------------------
qtm_connection = None
loop = asyncio.new_event_loop()
arduino = serial.Serial(port=SERIAL_PORT, baudrate=BAUD_RATE, timeout=1)
time.sleep(2)

event_log = []
press_times = {}
trial_number = 0
button_pool = []
current_button = None
num_buttons = 0

# Task management
tasks = []          # list of dicts: {name, uses_arduino}
selected_task = None
arduino_task_chosen = False

# ------------------ EVENT LOOP ------------------
def start_event_loop():
    asyncio.set_event_loop(loop)
    loop.run_forever()

threading.Thread(target=start_event_loop, daemon=True).start()

# ------------------ QTM CONTROL ------------------
async def start_qtm_recording():
    global qtm_connection
    try:
        qtm_connection = await qtm.connect("127.0.0.1")
        print("Connected to QTM.")
        await qtm_connection.take_control("")
        await qtm_connection.start()
        print("Recording started.")
    except Exception as e:
        print(f"Failed to start QTM recording: {e}")
        qtm_connection = None

async def stop_qtm_recording():
    global qtm_connection
    if qtm_connection:
        try:
            await qtm_connection.stop()
            print("Recording stopped.")
            qtm_connection.disconnect()
            print("Disconnected from QTM.")
            qtm_connection = None
        except Exception as e:
            print(f"Failed to stop QTM recording: {e}")
            def show_error():
                messagebox.showerror("Error", "Failed to stop recording.")
            root.after(0, show_error)

# ------------------ BEEP ------------------
def play_beep_blocking(task_name, uses_arduino):
    data, fs = sf.read(filename_beep, dtype='float32')
    event_log.append((trial_number, task_name, None, now_central().strftime('%H:%M:%S.%f')[:-3], "Beep Started", None, uses_arduino))
    sd.play(data, fs)
    sd.wait()

# ------------------ SERIAL READER ------------------
def read_serial():
    global trial_number, current_button
    while True:
        if arduino.in_waiting > 0:
            line = arduino.readline().decode(errors='ignore').strip()
            if not line:
                continue
            parts = line.split()
            if len(parts) != 2:
                continue

            event = parts[0]
            arduino_ms = int(parts[1])
            system_time = now_central().strftime('%H:%M:%S.%f')[:-3]

            if "_pressed" in event:
                button = event.split("_")[1]
                press_times[button] = arduino_ms
                event_text = f"#{button} - pressed"
                event_log.append((trial_number, selected_task["name"], current_button, system_time, event_text, None, selected_task["uses_arduino"]))
                print(f"[Trial {trial_number}] Button {current_button} [{system_time}] {event_text}")

                if current_button is not None:
                    arduino.write(f"LED_{current_button}_OFF\n".encode())

            elif "_released" in event:
                button = event.split("_")[1]
                if button in press_times:
                    duration = arduino_ms - press_times[button]
                    event_text = f"#{button} - released"
                    event_log.append((trial_number, selected_task["name"], current_button, system_time, event_text, duration, selected_task["uses_arduino"]))
                    print(f"[Trial {trial_number}] Button {current_button} [{system_time}] {event_text} (Duration: {duration} ms)")
                    del press_times[button]

serial_thread = threading.Thread(target=read_serial, daemon=True)
serial_thread.start()

# ------------------ TRIAL CONTROL ------------------
async def start_recording_and_trial():
    global trial_number, current_button, button_pool

    if selected_task is None:
        messagebox.showwarning("No Task Selected", "Please select a task before starting a trial.")
        return

    uses_arduino = selected_task["uses_arduino"]
    task_name = selected_task["name"]

    trial_number += 1

    event_log.append((trial_number, task_name, None, now_central().strftime('%H:%M:%S.%f')[:-3], "QTM Start Command Sent", None, uses_arduino))

    await start_qtm_recording()
    if qtm_connection is None:
        def show_error():
            messagebox.showerror("Error", "Failed to start recording.")
        root.after(0, show_error)
        return

    event_log.append((trial_number, task_name, None, now_central().strftime('%H:%M:%S.%f')[:-3], "QTM Recording Started", None, uses_arduino))

    await asyncio.sleep(0.5)

    beep_thread = threading.Thread(target=play_beep_blocking, args=(task_name, uses_arduino), daemon=True)
    beep_thread.start()

    if uses_arduino:
        if not button_pool:
            messagebox.showinfo("Done", "All buttons have been used.")
            return

        current_button = random.choice(button_pool)
        button_pool.remove(current_button)

        await asyncio.sleep(0.01)
        command = f"LED_{current_button}_ON\n"
        arduino.write(command.encode())
        arduino.flush()
        event_log.append((trial_number, task_name, current_button, now_central().strftime('%H:%M:%S.%f')[:-3], f"LED_{current_button}_Lit", None, uses_arduino))
        print(f"Trial {trial_number} ({task_name}): Beep played & Button {current_button} lit")
    else:
        print(f"Trial {trial_number} ({task_name}): Beep played (no Arduino)")

# GUI button actions
def on_start_button():
    asyncio.run_coroutine_threadsafe(start_recording_and_trial(), loop)

def on_stop_trial_button():
    asyncio.run_coroutine_threadsafe(stop_qtm_recording(), loop)

def on_end_button():
    export_to_excel()
    loop.call_soon_threadsafe(loop.stop)
    root.destroy()

# ------------------ EXCEL EXPORT ------------------
def export_to_excel():
    if not event_log:
        messagebox.showwarning("No Data", "No events to export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Trials"

    ws['A1'] = 'Trial'
    ws['B1'] = 'Task Name'
    ws['C1'] = 'Uses Arduino'
    ws['D1'] = 'Button Lit'
    ws['E1'] = 'Timestamp'
    ws['F1'] = 'Event'
    ws['G1'] = 'Duration (ms)'

    for idx, (trial, task_name, button, timestamp, event, duration, uses_arduino) in enumerate(event_log, start=2):
        ws[f"A{idx}"] = trial
        ws[f"B{idx}"] = task_name
        ws[f"C{idx}"] = "Yes" if uses_arduino else "No"
        ws[f"D{idx}"] = button
        ws[f"E{idx}"] = timestamp
        ws[f"F{idx}"] = event
        if duration is not None:
            ws[f"G{idx}"] = duration

    filename = f"trial_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    messagebox.showinfo("Export Successful", f"Saved as {filename}")

# ------------------ TASK SETUP ------------------
def setup_tasks():
    global tasks, arduino_task_chosen
    tasks = []
    for i in range(1, 9):
        name = simpledialog.askstring("Task Setup", f"Enter name for Task {i}:")
        if not name:
            name = f"Task {i}"

        uses_arduino = False
        if not arduino_task_chosen:
            answer = messagebox.askyesno("Arduino Question", f"Does task '{name}' use Arduino buttons?")
            if answer:
                uses_arduino = True
                arduino_task_chosen = True
        tasks.append({"name": name, "uses_arduino": uses_arduino})

# ------------------ GUI ------------------
def build_gui():
    global root, num_buttons, button_pool, selected_task

    root = tk.Tk()
    root.title("QTM + Seesaw Trial Controller")

    setup_tasks()

    # If Arduino task exists, ask how many buttons
    if any(t["uses_arduino"] for t in tasks):
        num_buttons = simpledialog.askinteger("Setup", "How many Seesaw buttons are connected? (1-4)", minvalue=1, maxvalue=4)
        button_pool = list(range(1, num_buttons + 1))

    # Dropdown to select task
    task_names = [t["name"] for t in tasks]
    selected_task_var = tk.StringVar(value=task_names[0])

    def update_selected_task(*args):
        global selected_task
        name = selected_task_var.get()
        for t in tasks:
            if t["name"] == name:
                selected_task = t
                break


    selected_task_var.trace("w", update_selected_task)
    update_selected_task()

    dropdown = ttk.OptionMenu(root, selected_task_var, task_names[0], *task_names)
    dropdown.pack(pady=10)

    start_btn = tk.Button(root, text="Start Trial", command=on_start_button, width=25, height=3)
    start_btn.pack(pady=10, padx=20)

    stop_btn = tk.Button(root, text="Stop Trial", command=on_stop_trial_button, width=25, height=3)
    stop_btn.pack(pady=10, padx=20)

    end_btn = tk.Button(root, text="End & Save", command=on_end_button, width=25, height=3)
    end_btn.pack(pady=10, padx=20)

    root.protocol("WM_DELETE_WINDOW", on_end_button)
    root.mainloop()

# ------------------ MAIN ------------------
if __name__ == "__main__":
    sync_windows_time()
    build_gui()
