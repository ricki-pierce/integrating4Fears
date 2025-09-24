"""
Integrated QTM + Arduino (Seesaw Buttons) + Excel Logger + Task Manager

When run, you will be asked to type in the names of the tasks (cup grabbing, buttoning shirt, light up LEDs, etc.).
For each task, you will be asked if it involves use of the Arduino. If you answer yes, the question will no longer appear for remaining tasks.
When you are ready to begin, you will see a dropdown menu at the top of the GUI. Make sure you select the correct task for whatever it is you are about to have the subject do. 
When you click start, QTM begins recording. 
When you click stop, QTM stops recording. 
Be sure to Save Recording after each trial, otherwise, you won't be able to reset QTM.
After saving, you need to reset so you can begin recording a new trial. 
The file naming is automatic. This is why you type in the tasks on the frontend. Trial numbers are also automatic. 
Once you are done recording ALL tasks and ALL trials, click End & Save. This will write an excel sheet with trial number, task name, if a button was lit, timestamps, and event description. 
"""

import asyncio
import threading
import serial
import time
import random
import numpy as np
import sounddevice as sd
import tkinter as tk
from tkinter import simpledialog, messagebox, ttk
from openpyxl import Workbook
from datetime import datetime
import qtm
import os
import soundfile as sf
import subprocess
import pytz

# ------------------ CONFIG ------------------
SERIAL_PORT = 'COM10'
BAUD_RATE = 115200
WORK_DIR = r"C:\\Users\\AoMV Lab\\ricki projects"
os.chdir(WORK_DIR)

filename_beep = r"C:\\Users\\AoMV Lab\\ricki projects\\500Hz500mstone.wav"

# ------------------ TIME SYNC ------------------

def sync_windows_time():
    try:
        output = subprocess.check_output(['w32tm', '/resync'], shell=True, stderr=subprocess.STDOUT, text=True)
        print(f"Time sync success:\n{output}")
    except subprocess.CalledProcessError as e:
        print(f"Time sync failed:\n{e.output}")

CENTRAL_TZ = pytz.timezone('America/Chicago')

def now_central():
    return datetime.now(CENTRAL_TZ)

# ------------------ GLOBALS ------------------
qtm_connection = None
loop = asyncio.new_event_loop()
arduino = serial.Serial(port=SERIAL_PORT, baudrate=BAUD_RATE, timeout=1)
time.sleep(2)

event_log = []
press_times = {}
button_pool = []
current_button = None
num_buttons = 0

# Task management
tasks = []
selected_task = None
arduino_task_chosen = False
# Per-task trial counter
task_trial_counts = {}

# Subject ID
subject_id = None

# ------------------ EVENT LOOP ------------------
def start_event_loop():
    asyncio.set_event_loop(loop)
    loop.run_forever()

threading.Thread(target=start_event_loop, daemon=True).start()

# ------------------ QTM CONTROL ------------------
async def start_qtm_recording():
    global qtm_connection
    try:
        if qtm_connection is None:
            qtm_connection = await qtm.connect("127.0.0.1")
            print("Connected to QTM.")
            await qtm_connection.take_control("")

        await qtm_connection.start()
        print("Recording started.")
    except Exception as e:
        print(f"Failed to start QTM recording: {e}")


async def stop_qtm_recording():
    global qtm_connection
    if qtm_connection:
        try:
            await qtm_connection.stop()
            print("Recording stopped.")
        except Exception as e:
            print(f"Failed to stop QTM recording: {e}")
            def show_error():
                messagebox.showerror("Error", "Failed to stop recording.")
            root.after(0, show_error)

# ------------------ QTM SAVE & RESET ------------------
async def save_qtm_recording():
    global qtm_connection
    if qtm_connection:
        try:
            task_name = selected_task["name"]
            trial_number = task_trial_counts[task_name]
            measurement_name = f"{task_name}_Trial{trial_number}_{subject_id}"
            await qtm_connection.save(measurement_name, overwrite=True)
            print(f"Recording saved as {measurement_name}")
        except Exception as e:
            print(f"Failed to save recording: {e}")
            def show_error():
                messagebox.showerror("Error", "Failed to save recording.")
            root.after(0, show_error)

async def reset_qtm():
    global qtm_connection
    if qtm_connection:
        try:
            await qtm_connection.new()
            print("QTM reset, ready for new measurement")
        except Exception as e:
            print(f"Failed to reset QTM: {e}")
            def show_error():
                messagebox.showerror("Error", "Failed to reset QTM.")
            root.after(0, show_error)
        qtm_connection.disconnect()
        print("Disconnected from QTM.")
        qtm_connection = None

# ------------------ BEEP ------------------
def play_beep_blocking(task_name, uses_arduino):
    trial_number = task_trial_counts[task_name]
    measurement_name = f"{task_name}_Trial{trial_number}"
    event_log.append((trial_number, task_name, None, now_central().strftime('%H:%M:%S.%f')[:-3], f"Beep Started", None, uses_arduino, None, subject_id))
    print(f"{measurement_name}: Beep Started")
    data, fs = sf.read(filename_beep, dtype='float32')
    sd.play(data, fs)
    sd.wait()
    
# ------------------ HAND ------------------

def ask_hand_used(trial_number, task_name):
    hand = messagebox.askquestion(
        "Hand Used", 
        f"For {task_name} (Trial {trial_number}), which hand did the subject use?\n\nOptions:\n- Left\n- Right\n- N/A",
        icon="question"
    )

    if hand == "yes":
        chosen_hand = "Left"
    else:
        response = messagebox.askyesnocancel("Hand Used", "Did the subject use the Right hand?\nYes = Right, No = N/A")
        if response is True:
            chosen_hand = "Right"
        elif response is False:
            chosen_hand = "N/A"
        else:
            chosen_hand = "N/A"

    timestamp = now_central().strftime('%H:%M:%S.%f')[:-3]

    # Note: added `hand_used` in position before 'duration'
    event_log.append((
        trial_number, 
        task_name, 
        None, 
        timestamp, 
        "Hand Used", 
        None, 
        selected_task["uses_arduino"], 
        chosen_hand, subject_id   # <--- NEW field
    ))

    print(f"{task_name}_Trial{trial_number}: Hand Used -> {chosen_hand}")



# ------------------ SERIAL READER ------------------
def read_serial():
    global current_button
    while True:
        if arduino.in_waiting > 0:
            line = arduino.readline().decode(errors='ignore').strip()
            if not line:
                continue
            parts = line.split()
            if len(parts) != 2:
                continue

            event = parts[0]
            arduino_ms = int(parts[1])
            system_time = now_central().strftime('%H:%M:%S.%f')[:-3]

            if "_pressed" in event:
                button = event.split("_")[1]
                press_times[button] = arduino_ms
                trial_number = task_trial_counts[selected_task['name']]
                measurement_name = f"{selected_task['name']}_Trial{trial_number}"
                event_text = f"#{button} - pressed"
                event_log.append((trial_number, selected_task["name"], current_button, system_time, event_text, None, selected_task["uses_arduino"], subject_id))
                print(f"{measurement_name}: Button {current_button} pressed")

                if current_button is not None:
                    arduino.write(f"LED_{current_button}_OFF\n".encode())

            elif "_released" in event:
                button = event.split("_")[1]
                if button in press_times:
                    duration = arduino_ms - press_times[button]
                    trial_number = task_trial_counts[selected_task['name']]
                    measurement_name = f"{selected_task['name']}_Trial{trial_number}"
                    event_text = f"#{button} - released"
                    event_log.append((trial_number, selected_task["name"], current_button, system_time, event_text, duration, selected_task["uses_arduino"], subject_id))
                    print(f"{measurement_name}: Button {current_button} released (Duration: {duration} ms)")
                    del press_times[button]

serial_thread = threading.Thread(target=read_serial, daemon=True)
serial_thread.start()

# ------------------ TRIAL CONTROL ------------------
async def start_recording_and_trial():
    global current_button, button_pool

    if selected_task is None:
        messagebox.showwarning("No Task Selected", "Please select a task before starting a trial.")
        return

    uses_arduino = selected_task["uses_arduino"]
    task_name = selected_task["name"]

    # Increment per-task counter
    task_trial_counts[task_name] += 1
    trial_number = task_trial_counts[task_name]
    measurement_name = f"{task_name}_Trial{trial_number}_{subject_id}"

    print(f"{measurement_name}: QTM Start Command Sent")
    event_log.append((trial_number, task_name, None, now_central().strftime('%H:%M:%S.%f')[:-3],
                      "QTM Start Command Sent", None, uses_arduino, subject_id))

    # --- Start QTM recording ---
    await start_qtm_recording()
    if qtm_connection is None:
        def show_error():
            messagebox.showerror("Error", "Failed to start recording.")
        root.after(0, show_error)
        return

    print(f"{measurement_name}: QTM Recording Started")
    event_log.append((trial_number, task_name, None, now_central().strftime('%H:%M:%S.%f')[:-3],
                      "QTM Recording Started", None, uses_arduino, subject_id))

    # --- Wait 500 ms after QTM actually started ---
    await asyncio.sleep(0.5)

    # --- Play beep + LED together ---
    def play_beep_and_led():
        # play new beep file
        data, fs = sf.read(r"C:\\Users\\AoMV Lab\\ricki projects\\500Hz500mstone.wav", dtype='float32')
        sd.play(data, fs)

        # Log beep
        event_log.append((trial_number, task_name, None, now_central().strftime('%H:%M:%S.%f')[:-3],
                          "Beep Started", None, uses_arduino, subject_id))
        print(f"{measurement_name}: Beep Started")

        # If Arduino task, light up LED at the same time
        if uses_arduino:
            if not button_pool:
                messagebox.showinfo("Done", "All buttons have been used.")
                return
            global current_button
            current_button = random.choice(button_pool)
            button_pool.remove(current_button)

            command = f"LED_{current_button}_ON\n"
            arduino.write(command.encode())
            arduino.flush()
            event_log.append((trial_number, task_name, current_button, now_central().strftime('%H:%M:%S.%f')[:-3],
                              f"LED_{current_button}_Lit", None, uses_arduino, subject_id))
            print(f"{measurement_name}: Button {current_button} lit")

    # Run beep + LED in a thread (so it doesn’t block)
    threading.Thread(target=play_beep_and_led, daemon=True).start()


# GUI button actions
def on_start_button():
    asyncio.run_coroutine_threadsafe(start_recording_and_trial(), loop)

def on_stop_trial_button():
    asyncio.run_coroutine_threadsafe(stop_qtm_recording(), loop)

    # Ask about hand used right after stopping
    if selected_task:
        trial_number = task_trial_counts[selected_task["name"]]
        task_name = selected_task["name"]
        root.after(100, lambda: ask_hand_used(trial_number, task_name))

def on_save_button():
    asyncio.run_coroutine_threadsafe(save_qtm_recording(), loop)

def on_reset_button():
    asyncio.run_coroutine_threadsafe(reset_qtm(), loop)


def on_end_button():
    export_to_excel()
    loop.call_soon_threadsafe(loop.stop)
    root.destroy()

# ------------------ EXCEL EXPORT ------------------
def export_to_excel():
    if not event_log:
        messagebox.showwarning("No Data", "No events to export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Trials"

    ws['A1'] = 'Trial'
    ws['B1'] = 'Task Name'
    ws['C1'] = 'Uses Arduino'
    ws['D1'] = 'Button Lit'
    ws['E1'] = 'Timestamp'
    ws['F1'] = 'Event'
    ws['G1'] = 'Duration (ms)'
    ws['H1'] = 'Hand Used'
    ws['I1'] = 'Subject ID'   # <--- NEW COLUMN


    for idx, entry in enumerate(event_log, start=2):
        # Old events may not have hand_used or subject_id
        if len(entry) == 7:
            trial, task_name, button, timestamp, event, duration, uses_arduino = entry
            hand_used = None
            sid = subject_id
        elif len(entry) == 8:
            trial, task_name, button, timestamp, event, duration, uses_arduino, hand_used = entry
            sid = subject_id
        else:
            trial, task_name, button, timestamp, event, duration, uses_arduino, hand_used, sid = entry
    
        ws[f"A{idx}"] = trial
        ws[f"B{idx}"] = task_name
        ws[f"C{idx}"] = "Yes" if uses_arduino else "No"
        ws[f"D{idx}"] = button
        ws[f"E{idx}"] = timestamp
        ws[f"F{idx}"] = event
        if duration is not None:
            ws[f"G{idx}"] = duration
        if hand_used is not None:
            ws[f"H{idx}"] = hand_used
        ws[f"I{idx}"] = sid


    filename = f"trial_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{subject_id}.xlsx"
    wb.save(filename)
    messagebox.showinfo("Export Successful", f"Saved as {filename}")


# ------------------ TASK SETUP ------------------
def setup_tasks():
    global tasks, arduino_task_chosen, task_trial_counts
    tasks = []
    task_trial_counts = {}

    # Ask user how many tasks they want
    num_tasks = simpledialog.askinteger(
        "Task Setup", 
        "How many tasks will be performed?", 
        minvalue=1, 
        maxvalue=50  # you can set a higher cap if needed
    )

    if not num_tasks:
        messagebox.showwarning("No Tasks", "No number entered. Defaulting to 1 task.")
        num_tasks = 1

    for i in range(1, num_tasks + 1):   # use the user’s number
        name = simpledialog.askstring("Task Setup", f"Enter name for Task {i}:")
        if not name:
            name = f"Task {i}"

        uses_arduino = False
        if not arduino_task_chosen:
            answer = messagebox.askyesno("Arduino Question", f"Does task '{name}' use Arduino buttons?")
            if answer:
                uses_arduino = True
                arduino_task_chosen = True

        tasks.append({"name": name, "uses_arduino": uses_arduino})
        task_trial_counts[name] = 0

# ------------------ GUI ------------------
def build_gui():
    global root, num_buttons, button_pool, selected_task, subject_id

    root = tk.Tk()
    root.title("QTM + Seesaw Trial Controller")

    # Ask for Subject ID first
    subject_id = simpledialog.askstring("Subject ID", "Enter Subject ID (alphanumeric):")
    if not subject_id:
        messagebox.showwarning("Missing ID", "No Subject ID entered. Using 'Unknown'.")
        subject_id = "Unknown"

    setup_tasks()

    if any(t["uses_arduino"] for t in tasks):
        num_buttons = simpledialog.askinteger("Setup", "How many Seesaw buttons are connected? (1-4)", minvalue=1, maxvalue=4)
        button_pool = list(range(1, num_buttons + 1))

    task_names = [t["name"] for t in tasks]
    selected_task_var = tk.StringVar(value=task_names[0])

    def update_selected_task(*args):
        global selected_task
        name = selected_task_var.get()
        for t in tasks:
            if t["name"] == name:
                selected_task = t
                break

    selected_task_var.trace("w", update_selected_task)
    update_selected_task()

    dropdown = ttk.OptionMenu(root, selected_task_var, task_names[0], *task_names)
    dropdown.pack(pady=10)

    start_btn = tk.Button(root, text="Start Trial", command=on_start_button, width=25, height=3)
    start_btn.pack(pady=10, padx=20)

    stop_btn = tk.Button(root, text="Stop Trial", command=on_stop_trial_button, width=25, height=3)
    stop_btn.pack(pady=10, padx=20)

    save_btn = tk.Button(root, text="Save Recording", command=on_save_button, width=25, height=3)
    save_btn.pack(pady=10, padx=20)

    reset_btn = tk.Button(root, text="Reset QTM", command=on_reset_button, width=25, height=3)
    reset_btn.pack(pady=10, padx=20)


    end_btn = tk.Button(root, text="End & Save", command=on_end_button, width=25, height=3)
    end_btn.pack(pady=10, padx=20)

    root.protocol("WM_DELETE_WINDOW", on_end_button)
    root.mainloop()

# ------------------ MAIN ------------------
if __name__ == "__main__":
    sync_windows_time()
    build_gui()
