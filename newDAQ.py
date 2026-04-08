"""
Integrated QTM + Neon Pupil Labs + Arduino (Seesaw Buttons) + Excel Logger
===========================================================================
Startup sequence per trial:
  1. estimate_time_offset measured BEFORE any recording starts (clean network)
  2. QTM starts recording → pc_time_qtm_start_ms captured immediately after
  3. 3-second delay
  4. Neon starts recording
  5. Beep + LED fire together to cue the subject

Metadata saved per trial (used by analysis code):
  - pc_time_qtm_start_ms       : PC clock (ms) at the moment QTM start was confirmed
  - phone_to_pc_time_offset_ms : averaged over 5 samples before recording starts
  - pc_time_neon_start_ms      : PC clock (ms) when Neon recording_start returned
                                 (informational only — not used in core analysis math)

File naming is automatic from subject ID + task name + trial number.
Arduino LED logic: one random button is lit per trial; pressing it turns it off.
After stopping each trial you are asked which hand the subject used.
Click End & Save when all trials are done to export the Excel log.
"""

import asyncio
import threading
import serial
import time
import json
import random
import numpy as np
import sounddevice as sd
import tkinter as tk
from tkinter import simpledialog, messagebox, ttk
from openpyxl import Workbook
from datetime import datetime
import qtm
import os
import soundfile as sf
import subprocess
import pytz
from pupil_labs.realtime_api.simple import discover_one_device

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

SERIAL_PORT   = 'COM6'
BAUD_RATE     = 115200
WORK_DIR      = r"C:\Users\AoMV Lab\ricki projects"
BEEP_FILE     = r"C:\Users\AoMV Lab\ricki projects\500Hz500mstone.wav"
NEON_DELAY_S  = 3.0        # seconds Neon starts AFTER QTM
N_OFFSET_SAMPLES = 5       # how many estimate_time_offset samples to average

os.chdir(WORK_DIR)

# ─────────────────────────────────────────────────────────────────────────────
# TIME HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def sync_windows_time():
    try:
        output = subprocess.check_output(
            ['w32tm', '/resync'], shell=True, stderr=subprocess.STDOUT, text=True
        )
        print(f"Time sync success:\n{output}")
    except subprocess.CalledProcessError as e:
        print(f"Time sync failed:\n{e.output}")

CENTRAL_TZ = pytz.timezone('America/Chicago')

def now_central():
    return datetime.now(CENTRAL_TZ)

def now_ms():
    """PC clock in milliseconds (Unix epoch). Used for all timing anchors."""
    return time.time() * 1000

# ─────────────────────────────────────────────────────────────────────────────
# GLOBALS
# ─────────────────────────────────────────────────────────────────────────────

neon_device     = None
qtm_connection  = None
loop            = asyncio.new_event_loop()
arduino         = serial.Serial(port=SERIAL_PORT, baudrate=BAUD_RATE, timeout=1)
time.sleep(2)   # let Arduino settle

event_log       = []
press_times     = {}
button_pool     = []
current_button  = None
num_buttons     = 0

tasks               = []
selected_task       = None
arduino_task_chosen = False
task_trial_counts   = {}
subject_id          = None

# ─────────────────────────────────────────────────────────────────────────────
# EVENT LOOP
# ─────────────────────────────────────────────────────────────────────────────

def start_event_loop():
    asyncio.set_event_loop(loop)
    loop.run_forever()

threading.Thread(target=start_event_loop, daemon=True).start()

# ─────────────────────────────────────────────────────────────────────────────
# QTM CONTROL
# ─────────────────────────────────────────────────────────────────────────────

async def start_qtm_recording():
    global qtm_connection
    try:
        if qtm_connection is None:
            qtm_connection = await qtm.connect("127.0.0.1")
            print("Connected to QTM.")
            await qtm_connection.take_control("")
        await qtm_connection.start()
        print("QTM recording started.")
    except Exception as e:
        print(f"Failed to start QTM recording: {e}")
        qtm_connection = None

async def stop_qtm_recording():
    global qtm_connection
    if qtm_connection:
        try:
            await qtm_connection.stop()
            print("QTM recording stopped.")
        except Exception as e:
            print(f"Failed to stop QTM recording: {e}")
            root.after(0, lambda: messagebox.showerror("Error", "Failed to stop QTM recording."))

async def save_qtm_recording():
    global qtm_connection
    if qtm_connection:
        try:
            task_name      = selected_task["name"]
            trial_number   = task_trial_counts[task_name]
            measurement_name = f"{task_name}_Trial{trial_number}_{subject_id}"
            await qtm_connection.save(measurement_name, overwrite=True)
            print(f"QTM recording saved as: {measurement_name}")
        except Exception as e:
            print(f"Failed to save QTM recording: {e}")
            root.after(0, lambda: messagebox.showerror("Error", "Failed to save QTM recording."))

async def reset_qtm():
    global qtm_connection
    if qtm_connection:
        try:
            await qtm_connection.new()
            print("QTM reset — ready for new measurement.")
        except Exception as e:
            print(f"Failed to reset QTM: {e}")
            root.after(0, lambda: messagebox.showerror("Error", "Failed to reset QTM."))
        qtm_connection.disconnect()
        qtm_connection = None
        print("Disconnected from QTM.")

# ─────────────────────────────────────────────────────────────────────────────
# NEON TIME OFFSET  (measured BEFORE recording starts for clean WiFi conditions)
# ─────────────────────────────────────────────────────────────────────────────

async def measure_time_offset():
    """
    Average N_OFFSET_SAMPLES estimates of the phone→PC clock offset.
    phone_to_pc_time_offset_ms = PC_time - Phone_time
    To convert a phone timestamp to PC time: phone_ts + offset
    """
    if neon_device is None:
        return None
    samples = []
    for i in range(N_OFFSET_SAMPLES):
        result = await asyncio.to_thread(neon_device.estimate_time_offset)
        samples.append(result.time_offset_ms.mean)
        print(f"  Offset sample {i+1}/{N_OFFSET_SAMPLES}: {result.time_offset_ms.mean:.2f} ms")
        await asyncio.sleep(0.2)
    averaged = float(np.mean(samples))
    print(f"  Averaged phone_to_pc_offset: {averaged:.2f} ms  (std: {np.std(samples):.2f} ms)")
    return averaged

# ─────────────────────────────────────────────────────────────────────────────
# MAIN TRIAL SEQUENCE
# ─────────────────────────────────────────────────────────────────────────────

async def start_recording_and_trial():
    global current_button, button_pool

    if selected_task is None:
        messagebox.showwarning("No Task Selected", "Please select a task before starting.")
        return

    uses_arduino = selected_task["uses_arduino"]
    task_name    = selected_task["name"]

    task_trial_counts[task_name] += 1
    trial_number     = task_trial_counts[task_name]
    measurement_name = f"{task_name}_Trial{trial_number}_{subject_id}"

    print(f"\n{'='*50}")
    print(f"Starting: {measurement_name}")
    print(f"{'='*50}")

    # ── 1. Measure time offset BEFORE any recording (clean network) ───────────
    print("Measuring phone-to-PC time offset...")
    phone_to_pc_time_offset_ms = await measure_time_offset()

    # ── 2. Start QTM and capture PC clock immediately after ───────────────────
    event_log.append((
        trial_number, task_name, None,
        now_central().strftime('%H:%M:%S.%f')[:-3],
        "QTM Start Command Sent", None, uses_arduino, None, subject_id
    ))

    await start_qtm_recording()

    if qtm_connection is None:
        root.after(0, lambda: messagebox.showerror("Error", "Failed to start QTM recording."))
        return

    # Capture PC clock the moment QTM confirms recording has started.
    # This is the anchor used in analysis: impact_time_pc_qtm = pc_time_qtm_start_ms + (frame * ms_per_frame)
    pc_time_qtm_start_ms = now_ms()
    print(f"QTM started. pc_time_qtm_start_ms = {pc_time_qtm_start_ms:.2f} ms")

    event_log.append((
        trial_number, task_name, None,
        now_central().strftime('%H:%M:%S.%f')[:-3],
        "QTM Recording Started", None, uses_arduino, None, subject_id
    ))

    # ── 3. Wait NEON_DELAY_S before starting Neon ─────────────────────────────
    print(f"Waiting {NEON_DELAY_S}s before starting Neon...")
    await asyncio.sleep(NEON_DELAY_S)

    # ── 4. Start Neon recording ───────────────────────────────────────────────
    pc_time_neon_start_ms = None
    if neon_device:
        try:
            await asyncio.to_thread(neon_device.recording_start)
            pc_time_neon_start_ms = now_ms()
            print(f"Neon started. pc_time_neon_start_ms = {pc_time_neon_start_ms:.2f} ms")
            print(f"Actual Neon delay from QTM start: {pc_time_neon_start_ms - pc_time_qtm_start_ms:.1f} ms")
        except Exception as e:
            print(f"Failed to start Neon recording: {e}")

    # ── 5. Save metadata JSON ─────────────────────────────────────────────────
    metadata = {
        "subject_id":               subject_id,
        "task_name":                task_name,
        "trial_number":             trial_number,
        "measurement_name":         measurement_name,
        # Core values used by analysis code:
        "pc_time_qtm_start_ms":     pc_time_qtm_start_ms,
        "phone_to_pc_time_offset_ms": phone_to_pc_time_offset_ms,
        # Informational (not used in core analysis math):
        "pc_time_neon_start_ms":    pc_time_neon_start_ms,
    }

    metadata_path = os.path.join(
        WORK_DIR,
        f"metadata_{task_name}_Trial{trial_number}_{subject_id}.json"
    )
    with open(metadata_path, "w") as f:
        json.dump(metadata, f, indent=4)
    print(f"Metadata saved: {metadata_path}")

    # ── 6. Wait a moment then fire beep + LED together to cue subject ─────────
    await asyncio.sleep(0.5)

    def play_beep_and_led():
        data, fs = sf.read(BEEP_FILE, dtype='float32')
        sd.play(data, fs)

        event_log.append((
            trial_number, task_name, None,
            now_central().strftime('%H:%M:%S.%f')[:-3],
            "Beep Started", None, uses_arduino, None, subject_id
        ))
        print(f"{measurement_name}: Beep fired.")

        if uses_arduino:
            if not button_pool:
                messagebox.showinfo("Done", "All buttons have been used.")
                return
            global current_button
            current_button = random.choice(button_pool)
            button_pool.remove(current_button)
            arduino.write(f"LED_{current_button}_ON\n".encode())
            arduino.flush()
            event_log.append((
                trial_number, task_name, current_button,
                now_central().strftime('%H:%M:%S.%f')[:-3],
                f"LED_{current_button}_Lit", None, uses_arduino, None, subject_id
            ))
            print(f"{measurement_name}: Button {current_button} lit.")

    threading.Thread(target=play_beep_and_led, daemon=True).start()


# ─────────────────────────────────────────────────────────────────────────────
# STOP TRIAL
# ─────────────────────────────────────────────────────────────────────────────

def on_stop_trial_button():

    async def stop_neon():
        if neon_device:
            try:
                await asyncio.to_thread(neon_device.recording_stop_and_save)
                print("Neon recording stopped and saved.")
            except Exception as e:
                print(f"Failed to stop Neon recording: {e}")
                root.after(0, lambda: messagebox.showerror("Error", f"Neon stop/save failed:\n{e}"))

    asyncio.run_coroutine_threadsafe(stop_qtm_recording(), loop)
    threading.Thread(target=lambda: asyncio.run(stop_neon()), daemon=True).start()

    if selected_task:
        trial_number = task_trial_counts[selected_task["name"]]
        task_name    = selected_task["name"]
        root.after(200, lambda: ask_hand_used(trial_number, task_name))


# ─────────────────────────────────────────────────────────────────────────────
# HAND USED DIALOG
# ─────────────────────────────────────────────────────────────────────────────

def ask_hand_used(trial_number, task_name):
    chosen_hand = "N/A"

    def set_hand(h):
        nonlocal chosen_hand
        chosen_hand = h
        hand_window.destroy()

    hand_window = tk.Toplevel(root)
    hand_window.title("Hand Used")
    hand_window.geometry("300x150")
    hand_window.grab_set()

    tk.Label(
        hand_window,
        text=f"{task_name} — Trial {trial_number}\nWhich hand did the subject use?",
        wraplength=280
    ).pack(pady=12)

    btn_frame = tk.Frame(hand_window)
    btn_frame.pack()
    for label in ("Left", "Right", "N/A"):
        tk.Button(
            btn_frame, text=label, width=10,
            command=lambda l=label: set_hand(l)
        ).pack(side=tk.LEFT, padx=5)

    hand_window.wait_window()

    event_log.append((
        trial_number, task_name, None,
        now_central().strftime('%H:%M:%S.%f')[:-3],
        "Hand Used", None, selected_task["uses_arduino"], chosen_hand, subject_id
    ))
    print(f"{task_name}_Trial{trial_number}: Hand Used → {chosen_hand}")


# ─────────────────────────────────────────────────────────────────────────────
# ARDUINO SERIAL READER
# ─────────────────────────────────────────────────────────────────────────────

def read_serial():
    global current_button
    while True:
        if arduino.in_waiting > 0:
            line = arduino.readline().decode(errors='ignore').strip()
            if not line:
                continue
            parts = line.split()
            if len(parts) != 2:
                continue

            event      = parts[0]
            arduino_ms = int(parts[1])
            sys_time   = now_central().strftime('%H:%M:%S.%f')[:-3]

            if selected_task is None:
                continue

            trial_number = task_trial_counts[selected_task['name']]

            if "_pressed" in event:
                button = event.split("_")[1]
                press_times[button] = arduino_ms
                event_log.append((
                    trial_number, selected_task["name"], current_button,
                    sys_time, f"#{button} - pressed", None,
                    selected_task["uses_arduino"], None, subject_id
                ))
                print(f"Trial {trial_number}: Button {current_button} pressed.")
                if current_button is not None:
                    arduino.write(f"LED_{current_button}_OFF\n".encode())

            elif "_released" in event:
                button = event.split("_")[1]
                if button in press_times:
                    duration = arduino_ms - press_times[button]
                    event_log.append((
                        trial_number, selected_task["name"], current_button,
                        sys_time, f"#{button} - released", duration,
                        selected_task["uses_arduino"], None, subject_id
                    ))
                    print(f"Trial {trial_number}: Button {current_button} released ({duration} ms).")
                    del press_times[button]

threading.Thread(target=read_serial, daemon=True).start()


# ─────────────────────────────────────────────────────────────────────────────
# GUI BUTTON CALLBACKS
# ─────────────────────────────────────────────────────────────────────────────

def on_start_button():
    asyncio.run_coroutine_threadsafe(start_recording_and_trial(), loop)

def on_save_button():
    asyncio.run_coroutine_threadsafe(save_qtm_recording(), loop)

def on_reset_button():
    asyncio.run_coroutine_threadsafe(reset_qtm(), loop)

def on_end_button():
    export_to_excel()
    loop.call_soon_threadsafe(loop.stop)
    root.destroy()


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_to_excel():
    if not event_log:
        messagebox.showwarning("No Data", "No events to export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Trials"

    headers = [
        "Subject ID", "Task Name", "Trial", "Timestamp",
        "Event", "Hand Used", "Uses Arduino", "Button Lit", "Duration (ms)"
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, entry in enumerate(event_log, start=2):
        trial, task_name, button, timestamp, event, duration, uses_arduino, hand_used, sid = entry
        ws[f"A{row_idx}"] = sid
        ws[f"B{row_idx}"] = task_name
        ws[f"C{row_idx}"] = trial
        ws[f"D{row_idx}"] = timestamp
        ws[f"E{row_idx}"] = event
        ws[f"F{row_idx}"] = hand_used
        ws[f"G{row_idx}"] = "Yes" if uses_arduino else "No"
        ws[f"H{row_idx}"] = button
        ws[f"I{row_idx}"] = duration

    filename = f"trial_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{subject_id}.xlsx"
    wb.save(filename)
    messagebox.showinfo("Export Successful", f"Saved as {filename}")
    print(f"Excel log saved: {filename}")


# ─────────────────────────────────────────────────────────────────────────────
# TASK SETUP
# ─────────────────────────────────────────────────────────────────────────────

def setup_tasks():
    global tasks, arduino_task_chosen, task_trial_counts
    tasks             = []
    task_trial_counts = {}

    num_tasks = simpledialog.askinteger(
        "Task Setup", "How many tasks will be performed?", minvalue=1, maxvalue=50
    )
    if not num_tasks:
        messagebox.showwarning("No Tasks", "Defaulting to 1 task.")
        num_tasks = 1

    for i in range(1, num_tasks + 1):
        name = simpledialog.askstring("Task Setup", f"Enter name for Task {i}:")
        if not name:
            name = f"Task{i}"

        uses_arduino = False
        if not arduino_task_chosen:
            if messagebox.askyesno("Arduino", f"Does '{name}' use Arduino buttons?"):
                uses_arduino        = True
                arduino_task_chosen = True

        tasks.append({"name": name, "uses_arduino": uses_arduino})
        task_trial_counts[name] = 0


# ─────────────────────────────────────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────────────────────────────────────

def build_gui():
    global root, num_buttons, button_pool, selected_task, subject_id

    root = tk.Tk()
    root.title("QTM + Neon Trial Controller")

    # Try to connect to Neon in background so GUI doesn't freeze
    def connect_neon():
        global neon_device
        neon_device = discover_one_device(max_search_duration_seconds=10)
        if neon_device:
            print("Neon device connected.")
        else:
            print("No Neon device found — continuing without eye tracking.")
    threading.Thread(target=connect_neon, daemon=True).start()

    # Subject ID
    subject_id = simpledialog.askstring("Subject ID", "Enter Subject ID:")
    if not subject_id:
        messagebox.showwarning("Missing ID", "No Subject ID entered. Using 'Unknown'.")
        subject_id = "Unknown"

    setup_tasks()

    if any(t["uses_arduino"] for t in tasks):
        num_buttons = simpledialog.askinteger(
            "Setup", "How many Seesaw buttons are connected? (1–4)",
            minvalue=1, maxvalue=4
        )
        button_pool = list(range(1, num_buttons + 1))

    task_names        = [t["name"] for t in tasks]
    selected_task_var = tk.StringVar(value=task_names[0])

    def update_selected_task(*_):
        global selected_task
        name = selected_task_var.get()
        selected_task = next(t for t in tasks if t["name"] == name)

    selected_task_var.trace("w", update_selected_task)
    update_selected_task()

    # Layout
    ttk.OptionMenu(root, selected_task_var, task_names[0], *task_names).pack(pady=10)

    for label, cmd in [
        ("Start Trial",     on_start_button),
        ("Stop Trial",      on_stop_trial_button),
        ("Save Recording",  on_save_button),
        ("Reset QTM",       on_reset_button),
        ("End & Save",      on_end_button),
    ]:
        tk.Button(root, text=label, command=cmd, width=25, height=3).pack(pady=6, padx=20)

    root.protocol("WM_DELETE_WINDOW", on_end_button)
    root.mainloop()


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    sync_windows_time()
    build_gui()
