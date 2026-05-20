"""

*using this code on 5/17/26, i ran 25 trials (20 useable) and got that there was a delay of about 20.9ms +/- 2.6ms regarding moment of impact.

Integrated QTM + Arduino (Seesaw Buttons) + Excel Logger + Task Manager

When run, you will be asked to type in the names of the tasks (cup grabbing, buttoning shirt, light up LEDs, etc.).
For each task, you will be asked if it involves use of the Arduino. If you answer yes, the question will no longer appear for remaining tasks.
When you are ready to begin, you will see a dropdown menu at the top of the GUI. Make sure you select the correct task for whatever it is you are about to have the subject do. 
When you click start, QTM begins recording AND SO DO THE NEON GLASSES. 
When you click stop, QTM stops recording AND SO DO THE NEON GLASSES. 
Be sure to Save Recording after each trial, otherwise, you won't be able to reset QTM.
After saving, you need to reset so you can begin recording a new trial. 
The file naming is automatic. This is why you type in the tasks on the frontend. Trial numbers are also automatic. (Note: the file naming does not extend to the videos in the Neon folder.)
Once you are done recording ALL tasks and ALL trials, click End & Save. This will write an excel sheet with trial number, task name, if a button was lit, timestamps, and event description. 
"""

import asyncio
import threading
import serial
import time
import json
import random
import numpy as np
import sounddevice as sd
import tkinter as tk
from tkinter import simpledialog, messagebox, ttk
from openpyxl import Workbook
from datetime import datetime
import qtm
import os
import soundfile as sf
import subprocess
import pytz
from pupil_labs.realtime_api.simple import discover_one_device
neon_device = None
# ------------------ CONFIG ------------------
SERIAL_PORT = 'COM6'
BAUD_RATE = 115200
WORK_DIR = r"C:\\Users\\AoMV Lab\\ricki projects"
os.chdir(WORK_DIR)
filename_beep = r"C:\\Users\\AoMV Lab\\ricki projects\\500Hz500mstone.wav"

# ------------------ TIME SYNC ------------------

def sync_windows_time():
    try:
        output = subprocess.check_output(['w32tm', '/resync'], shell=True, stderr=subprocess.STDOUT, text=True)
        print(f"Time sync success:\n{output}")
    except subprocess.CalledProcessError as e:
        print(f"Time sync failed:\n{e.output}")

CENTRAL_TZ = pytz.timezone('America/Chicago')

def now_central():
    return datetime.now(CENTRAL_TZ)

# ------------------ GLOBALS ------------------
qtm_connection = None
loop = asyncio.new_event_loop()
arduino = serial.Serial(port=SERIAL_PORT, baudrate=BAUD_RATE, timeout=1)
time.sleep(2)
serial_lock = threading.Lock()

event_log = []
press_times = {}
button_pool = []
current_button = None
num_buttons = 0

# Task management
tasks = []
selected_task = None
arduino_task_chosen = False
# Per-task trial counter
task_trial_counts = {}

# Subject ID
subject_id = None

# ------------------ EVENT LOOP ------------------
def start_event_loop():
    asyncio.set_event_loop(loop)
    loop.run_forever()

threading.Thread(target=start_event_loop, daemon=True).start()

# ------------------ QTM CONTROL ------------------
async def start_qtm_recording():
    global qtm_connection
    try:
        if qtm_connection is None:
            qtm_connection = await qtm.connect("127.0.0.1")
            print("Connected to QTM.")
            #print(dir(qtm_connection))
            await qtm_connection.take_control("")

        await qtm_connection.start()
        print("Recording started.")
    except Exception as e:
        print(f"Failed to start QTM recording: {e}")


async def stop_qtm_recording():
    global qtm_connection
    if qtm_connection:
        try:
            await qtm_connection.stop()
            print("Recording stopped.")
        except Exception as e:
            print(f"Failed to stop QTM recording: {e}")
            def show_error():
                messagebox.showerror("Error", "Failed to stop recording.")
            root.after(0, show_error)

# ------------------ TEST TRIGGER ------------------

async def send_test_trigger():
    global qtm_connection, neon_device

    timestamp = now_central().strftime('%H:%M:%S.%f')[:-3]

    print(f"TEST TRIGGER fired at {timestamp}")

    # Log locally
    event_log.append((
        task_trial_counts[selected_task["name"]],
        selected_task["name"],
        None,
        timestamp,
        "TEST_TRIGGER",
        None,
        selected_task["uses_arduino"],
        None,
        subject_id
    ))

    # ---------------- QTM MARKER ----------------
    if qtm_connection:
        try:
            # Create QTM event marker
            await qtm_connection.set_qtm_event("TEST_TRIGGER")

            print("QTM trigger sent")

        except Exception as e:
            print(f"Failed to send QTM trigger: {e}")

    # ---------------- PUPIL LABS MARKER ----------------
    if neon_device:
        try:
            await asyncio.to_thread(
                neon_device.send_event,
                "TEST_TRIGGER"
            )

            print("Neon trigger sent")

        except Exception as e:
            print(f"Failed to send Neon trigger: {e}")

# ------------------ QTM SAVE & RESET ------------------
async def save_qtm_recording():
    global qtm_connection
    if qtm_connection:
        try:
            task_name = selected_task["name"]
            trial_number = task_trial_counts[task_name]
            measurement_name = f"{task_name}_Trial{trial_number}_{subject_id}"
            await qtm_connection.save(measurement_name, overwrite=True)
            print(f"Recording saved as {measurement_name}")
        except Exception as e:
            print(f"Failed to save recording: {e}")
            def show_error():
                messagebox.showerror("Error", "Failed to save recording.")
            root.after(0, show_error)

async def reset_qtm():
    global qtm_connection
    if qtm_connection:
        try:
            await qtm_connection.new()
            print("QTM reset, ready for new measurement")
        except Exception as e:
            print(f"Failed to reset QTM: {e}")
            def show_error():
                messagebox.showerror("Error", "Failed to reset QTM.")
            root.after(0, show_error)
        qtm_connection.disconnect()
        print("Disconnected from QTM.")
        qtm_connection = None

# ------------------ BEEP ------------------
def play_beep_blocking(task_name, uses_arduino):
    trial_number = task_trial_counts[task_name]
    measurement_name = f"{task_name}_Trial{trial_number}"
    event_log.append((
        trial_number,
        task_name,
        None,  # button
        now_central().strftime('%H:%M:%S.%f')[:-3],
        "Beep Started",
        None,  # duration
        uses_arduino,
        None,  # hand_used placeholder
        subject_id
    ))

    print(f"{measurement_name}: Beep Started")
    data, fs = sf.read(filename_beep, dtype='float32')
    sd.play(data, fs)
    sd.wait()
    
# ------------------ HAND ------------------

def ask_hand_used(trial_number, task_name):
    def set_hand(hand_choice):
        nonlocal chosen_hand
        chosen_hand = hand_choice
        hand_window.destroy()

    chosen_hand = "N/A"  # default
    hand_window = tk.Toplevel(root)
    hand_window.title("Hand Used")
    hand_window.geometry("300x150")
    hand_window.grab_set()  # Make it modal (block interaction with main window)

    label = tk.Label(
        hand_window,
        text=f"For {task_name} (Trial {trial_number}), which hand did the subject use?",
        wraplength=280
    )
    label.pack(pady=10)

    btn_frame = tk.Frame(hand_window)
    btn_frame.pack(pady=10)

    tk.Button(btn_frame, text="Left", width=10, command=lambda: set_hand("Left")).grid(row=0, column=0, padx=5)
    tk.Button(btn_frame, text="Right", width=10, command=lambda: set_hand("Right")).grid(row=0, column=1, padx=5)
    tk.Button(btn_frame, text="N/A", width=10, command=lambda: set_hand("N/A")).grid(row=0, column=2, padx=5)

    hand_window.wait_window()  # Wait until closed

    timestamp = now_central().strftime('%H:%M:%S.%f')[:-3]
    event_log.append((
        trial_number,
        task_name,
        None,
        timestamp,
        "Hand Used",
        None,
        selected_task["uses_arduino"],
        chosen_hand,
        subject_id
    ))

    print(f"{task_name}_Trial{trial_number}: Hand Used -> {chosen_hand}")



# ------------------ SERIAL READER ------------------
def read_serial():
    global current_button
    while True:
        if arduino.in_waiting > 0:
            line = arduino.readline().decode(errors='ignore').strip()
            if not line:
                continue
            parts = line.split()
            if len(parts) != 2:
                continue

            event = parts[0]
            arduino_ms = int(parts[1])
            system_time = now_central().strftime('%H:%M:%S.%f')[:-3]

            if "_pressed" in event:
                button = event.split("_")[1]
                press_times[button] = arduino_ms
                trial_number = task_trial_counts[selected_task['name']]
                measurement_name = f"{selected_task['name']}_Trial{trial_number}"
                event_text = f"#{button} - pressed"
                event_log.append((trial_number, selected_task["name"], current_button, system_time, event_text, None, selected_task["uses_arduino"], None, subject_id))
                print(f"{measurement_name}: Button {current_button} pressed")

                if current_button is not None:
                    with serial_lock:
                        arduino.write(f"LED_{current_button}_OFF\n".encode())
            elif "_released" in event:
                button = event.split("_")[1]
                if button in press_times:
                    duration = arduino_ms - press_times[button]
                    trial_number = task_trial_counts[selected_task['name']]
                    measurement_name = f"{selected_task['name']}_Trial{trial_number}"
                    event_text = f"#{button} - released"
                    event_log.append((trial_number, selected_task["name"], current_button, system_time, event_text, duration, selected_task["uses_arduino"], None, subject_id))
                    print(f"{measurement_name}: Button {current_button} released (Duration: {duration} ms)")
                    del press_times[button]

serial_thread = threading.Thread(target=read_serial, daemon=True)
serial_thread.start()

# ------------------ TRIAL CONTROL ------------------
async def start_recording_and_trial():
    global current_button, button_pool

    if selected_task is None:
        messagebox.showwarning("No Task Selected", "Please select a task before starting a trial.")
        return

    uses_arduino = selected_task["uses_arduino"]
    task_name = selected_task["name"]

    # Increment per-task counter
    task_trial_counts[task_name] += 1
    trial_number = task_trial_counts[task_name]
    measurement_name = f"{task_name}_Trial{trial_number}_{subject_id}"

    print(f"{measurement_name}: QTM Start Command Sent")
    event_log.append((
        trial_number,
        task_name,
        None,  # button
        now_central().strftime('%H:%M:%S.%f')[:-3],
        "QTM Start Command Sent",
        None,  # duration
        uses_arduino,
        None,  # hand_used placeholder
        subject_id
    ))

    # --- Start Neon FIRST (so sensor boots before QTM begins capturing) ---
    phone_to_pc_time_offset_ms = None  # default if Neon not connected
    
    if neon_device:
            try:
                await asyncio.to_thread(neon_device.recording_start)
                print("Neon recording started")
    
                try:
                    offset_result = await asyncio.to_thread(neon_device.estimate_time_offset)
                    phone_to_pc_time_offset_ms = offset_result.time_offset_ms.mean
                    print(f"Phone-to-PC time offset: {phone_to_pc_time_offset_ms} ms")
                except Exception as e:
                    print(f"Could not get time offset: {e}")
                    phone_to_pc_time_offset_ms = None
    
            except Exception as e:
                print(f"Failed to start Neon recording: {e}")
    
        # --- Give Neon time to finish booting its sensor before QTM starts ---
    await asyncio.sleep(1.5)
    
        # --- Now start QTM ---
    await start_qtm_recording()
    
    if qtm_connection is None:
            def show_error():
                messagebox.showerror("Error", "Failed to start recording.")
            root.after(0, show_error)
            return
    
        # Capture QTM start time IMMEDIATELY after start confirmation
    pc_time_qtm_start_ms = time.time() * 1000
    print(f"QTM start time captured: {pc_time_qtm_start_ms}")
    
        # --- Write metadata.json for this trial ---
    metadata = {
            "subject_id": subject_id,
            "task_name": task_name,
            "trial_number": trial_number,
            "pc_time_qtm_start_ms": pc_time_qtm_start_ms,
            "phone_to_pc_time_offset_ms": phone_to_pc_time_offset_ms,
            "measurement_name": measurement_name
        }
    
    metadata_filename = os.path.join(
            WORK_DIR,
            f"metadata_{task_name}_Trial{trial_number}_{subject_id}.json"
        )
    
    with open(metadata_filename, "w") as f:
            json.dump(metadata, f, indent=4)
    
    print(f"Metadata saved: {metadata_filename}")

    print(f"{measurement_name}: QTM Recording Started")

        # --- Start Neon recording ---

    event_log.append((
        trial_number,
        task_name,
        None,  # button
        now_central().strftime('%H:%M:%S.%f')[:-3],
        "QTM Recording Started",
        None,  # duration
        uses_arduino,
        None,  # hand_used placeholder
        subject_id
    ))


    # --- Wait 500 ms after QTM actually started ---
    await asyncio.sleep(1.5)

    # --- Play beep + LED together ---
    def play_beep_and_led():
        # play new beep file
        data, fs = sf.read(r"C:\\Users\\AoMV Lab\\ricki projects\\500Hz500mstone.wav", dtype='float32')
        sd.play(data, fs)

        # Log beep
        event_log.append((trial_number, task_name, None, now_central().strftime('%H:%M:%S.%f')[:-3],
                          "Beep Started", None, uses_arduino, None, subject_id))
        print(f"{measurement_name}: Beep Started")

        # If Arduino task, light up LED at the same time
        if uses_arduino:
            if not button_pool:
                messagebox.showinfo("Done", "All buttons have been used.")
                return
            global current_button
            current_button = random.choice(button_pool)
            button_pool.remove(current_button)

            command = f"LED_{current_button}_ON\n"
            with serial_lock:
                arduino.write(command.encode())
                arduino.flush()
            event_log.append((trial_number, task_name, current_button, now_central().strftime('%H:%M:%S.%f')[:-3],
                              f"LED_{current_button}_Lit", None, uses_arduino, None, subject_id))

            print(f"{measurement_name}: Button {current_button} lit")

    # Run beep + LED in a thread (so it doesn’t block)
    threading.Thread(target=play_beep_and_led, daemon=True).start()




# GUI button actions
def on_start_button():
    asyncio.run_coroutine_threadsafe(start_recording_and_trial(), loop)

def on_stop_trial_button():
    
    async def stop_neon_recording():
        if neon_device:
            try:
                await asyncio.to_thread(neon_device.recording_stop_and_save)
                print("Neon recording stopped and saved")
            except Exception as e:
                print(f"Failed to stop Neon recording: {e}")
                def show_error():
                    messagebox.showerror("Error", f"Neon stop/save failed:\n{e}")
                root.after(0, show_error)

    asyncio.run_coroutine_threadsafe(stop_qtm_recording(), loop)
    asyncio.run_coroutine_threadsafe(stop_neon_recording(), loop)


    # Ask about hand used right after stopping
    if selected_task:
        trial_number = task_trial_counts[selected_task["name"]]
        task_name = selected_task["name"]
        root.after(100, lambda: ask_hand_used(trial_number, task_name))

def on_save_button():
    asyncio.run_coroutine_threadsafe(save_qtm_recording(), loop)

def on_reset_button():
    asyncio.run_coroutine_threadsafe(reset_qtm(), loop)


def on_end_button():
    export_to_excel()
    loop.call_soon_threadsafe(loop.stop)
    root.destroy()

def on_trigger_button():
    asyncio.run_coroutine_threadsafe(send_test_trigger(), loop)
# ------------------ EXCEL EXPORT ------------------
def export_to_excel():
    if not event_log:
        messagebox.showwarning("No Data", "No events to export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Trials"

    ws['A1'] = 'Subject ID'
    ws['B1'] = 'Task Name'
    ws['C1'] = 'Trial'
    ws['D1'] = 'Timestamp'
    ws['E1'] = 'Event'
    ws['F1'] = 'Hand Used'
    ws['G1'] = 'Uses Arduino'
    ws['H1'] = 'Button Lit'
    ws['I1'] = 'Duration (ms)'


    for idx, entry in enumerate(event_log, start=2):
            # unpack the 9-element tuple
            trial, task_name, button, timestamp, event, duration, uses_arduino, hand_used, sid = entry

            ws[f"A{idx}"] = sid
            ws[f"B{idx}"] = task_name
            ws[f"C{idx}"] = trial
            ws[f"D{idx}"] = timestamp
            ws[f"E{idx}"] = event
            ws[f"F{idx}"] = hand_used
            ws[f"G{idx}"] = "Yes" if uses_arduino else "No"
            ws[f"H{idx}"] = button
            ws[f"I{idx}"] = duration

    filename = f"trial_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{subject_id}.xlsx"
    wb.save(filename)
    messagebox.showinfo("Export Successful", f"Saved as {filename}")


# ------------------ TASK SETUP ------------------
def setup_tasks():
    global tasks, arduino_task_chosen, task_trial_counts
    tasks = []
    task_trial_counts = {}

    # Ask user how many tasks they want
    num_tasks = simpledialog.askinteger(
        "Task Setup", 
        "How many tasks will be performed?", 
        minvalue=1, 
        maxvalue=50  # you can set a higher cap if needed
    )

    if not num_tasks:
        messagebox.showwarning("No Tasks", "No number entered. Defaulting to 1 task.")
        num_tasks = 1

    for i in range(1, num_tasks + 1):   # use the user’s number
        name = simpledialog.askstring("Task Setup", f"Enter name for Task {i}:")
        if not name:
            name = f"Task {i}"

        uses_arduino = False
        if not arduino_task_chosen:
            answer = messagebox.askyesno("Arduino Question", f"Does task '{name}' use Arduino buttons?")
            if answer:
                uses_arduino = True
                arduino_task_chosen = True

        tasks.append({"name": name, "uses_arduino": uses_arduino})
        task_trial_counts[name] = 0

# ------------------ GUI ------------------
def build_gui():

    global root, num_buttons, button_pool, selected_task, subject_id

    root = tk.Tk()
    root.title("QTM + Seesaw Trial Controller")

    global neon_device
    try:
        # Try to discover Neon device in a thread so GUI doesn't freeze
        def connect_neon():
            global neon_device
            neon_device = discover_one_device(max_search_duration_seconds=10)
            if neon_device:
                print("Neon device connected")
                print(dir(neon_device))  # ✅ ADD THIS LINE TEMPORARILY
            else:
                print("No Neon device found")

        threading.Thread(target=connect_neon, daemon=True).start()
    except Exception as e:
        print(f"Neon connection error: {e}")


    # Ask for Subject ID first
    subject_id = simpledialog.askstring("Subject ID", "Enter Subject ID (alphanumeric):")
    if not subject_id:
        messagebox.showwarning("Missing ID", "No Subject ID entered. Using 'Unknown'.")
        subject_id = "Unknown"

    setup_tasks()

    if any(t["uses_arduino"] for t in tasks):
        num_buttons = simpledialog.askinteger("Setup", "How many Seesaw buttons are connected? (1-4)", minvalue=1, maxvalue=4)
        button_pool = list(range(1, num_buttons + 1))

    task_names = [t["name"] for t in tasks]
    selected_task_var = tk.StringVar(value=task_names[0])

    def update_selected_task(*args):
        global selected_task
        name = selected_task_var.get()
        for t in tasks:
            if t["name"] == name:
                selected_task = t
                break


    selected_task_var.trace("w", update_selected_task)
    update_selected_task()

    dropdown = ttk.OptionMenu(root, selected_task_var, task_names[0], *task_names)
    dropdown.pack(pady=10)

    start_btn = tk.Button(root, text="Start Trial", command=on_start_button, width=25, height=3)
    start_btn.pack(pady=10, padx=20)

    stop_btn = tk.Button(root, text="Stop Trial", command=on_stop_trial_button, width=25, height=3)
    stop_btn.pack(pady=10, padx=20)

    trigger_btn = tk.Button(
        root,
        text="Test Trigger",
        command=on_trigger_button,
        width=25,
        height=3,
        bg="red",
        fg="white"
    )
    trigger_btn.pack(pady=10, padx=20)

    save_btn = tk.Button(root, text="Save Recording", command=on_save_button, width=25, height=3)
    save_btn.pack(pady=10, padx=20)

    reset_btn = tk.Button(root, text="Reset QTM", command=on_reset_button, width=25, height=3)
    reset_btn.pack(pady=10, padx=20)


    end_btn = tk.Button(root, text="End & Save", command=on_end_button, width=25, height=3)
    end_btn.pack(pady=10, padx=20)

    root.protocol("WM_DELETE_WINDOW", on_end_button)
    root.mainloop()

# ------------------ MAIN ------------------
if __name__ == "__main__":
    sync_windows_time()
    build_gui()
