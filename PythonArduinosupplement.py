# --- Integrated QTM + Arduino + Excel Logger ---

import asyncio
import threading
import serial
import time
import random
import numpy as np
import sounddevice as sd
import tkinter as tk
from tkinter import simpledialog, messagebox
from openpyxl import Workbook
from datetime import datetime
import qtm
import os
import soundfile as sf

# ------------------ CONFIG ------------------
# Adjust to match your system
SERIAL_PORT = 'COM10'    # Change if Arduino is on different port
BAUD_RATE = 9600
WORK_DIR = r"C:\Users\AoMV Lab\ricki projects"
os.chdir(WORK_DIR)

# ------------------ GLOBALS ------------------
qtm_connection = None
loop = asyncio.new_event_loop()

arduino = serial.Serial(port=SERIAL_PORT, baudrate=BAUD_RATE, timeout=1)
time.sleep(2)  # wait for Arduino reset

event_log = []     # list of (trial, led, system_time, event, duration)
press_times = {}   # track button press times
trial_number = 0
led_pool = []      # LEDs available for random choice
current_led = None
num_leds = 0

# ------------------ EVENT LOOP ------------------
def start_event_loop():
    asyncio.set_event_loop(loop)
    loop.run_forever()

threading.Thread(target=start_event_loop, daemon=True).start()

# ------------------ QTM CONTROL ------------------
async def start_qtm_recording():
    global qtm_connection
    try:
        qtm_connection = await qtm.connect("127.0.0.1")
        print("Connected to QTM.")
        await qtm_connection.take_control("")
        await qtm_connection.start()
        print("Recording started.")
    except Exception as e:
        print(f"Failed to start QTM recording: {e}")
        qtm_connection = None

async def stop_qtm_recording():
    global qtm_connection
    if qtm_connection:
        try:
            await qtm_connection.stop()
            print("Recording stopped.")
            qtm_connection.disconnect()
            print("Disconnected from QTM.")
            qtm_connection = None
        except Exception as e:
            print(f"Failed to stop QTM recording: {e}")
            def show_error():
                messagebox.showerror("Error", "Failed to stop recording.")
            root.after(0, show_error)

# ------------------ BEEP ------------------
def play_beep_blocking():
    # Replace this path with the full path to your .wav file
    filename = r"C:\Users\AoMV Lab\ricki projects\Silenceplus500hz1000mstone.wav"
    
    data, fs = sf.read(filename, dtype='float32')  # Read WAV file
    sd.play(data, fs)                              # Play sound
    sd.wait()                                      # Wait until finished

# ------------------ SERIAL READER ------------------
def read_serial():
    global trial_number, current_led
    while True:
        if arduino.in_waiting > 0:
            line = arduino.readline().decode().strip()
            if not line:
                continue

            parts = line.split()
            if len(parts) != 2:
                continue

            event = parts[0]
            arduino_ms = int(parts[1])
            system_time = datetime.now().strftime('%H:%M:%S.%f')[:-3]

            if "_pressed" in event:
                button = event.split("_")[1]
                press_times[button] = arduino_ms
                event_text = f"#{button} - pressed"
                event_log.append((trial_number, current_led, system_time, event_text, None))
                print(f"[Trial {trial_number}] LED {current_led} [{system_time}] {event_text}")

                # Turn LED off as soon as button is pressed
                if current_led is not None:
                    arduino.write(f"LED_{current_led}_OFF\n".encode())

            elif "_released" in event:
                button = event.split("_")[1]
                if button in press_times:
                    duration = arduino_ms - press_times[button]
                    event_text = f"#{button} - released"
                    event_log.append((trial_number, current_led, system_time, event_text, duration))
                    print(f"[Trial {trial_number}] LED {current_led} [{system_time}] {event_text} (Duration: {duration} ms)")
                    del press_times[button]

serial_thread = threading.Thread(target=read_serial, daemon=True)
serial_thread.start()

# ------------------ TRIAL CONTROL ------------------
async def start_recording_and_trial():
    global trial_number, current_led, led_pool

    if not led_pool:
        messagebox.showinfo("Done", "All LEDs have been used.")
        return

    trial_number += 1
    await start_qtm_recording()
    if qtm_connection is None:
        def show_error():
            messagebox.showerror("Error", "Failed to start recording.")
        root.after(0, show_error)
        return

    await asyncio.sleep(2)

    # --- Choose LED first ---
    current_led = random.choice(led_pool)
    led_pool.remove(current_led)

# --- Start beep in a separate thread ---
    beep_thread = threading.Thread(target=play_beep_blocking, daemon=True)
    beep_thread.start()

    # --- Wait 500 ms before turning on LED ---
    await asyncio.sleep(1.0)
    command = f"LED_{current_led}_ON\n"
    arduino.write(command.encode())
    arduino.flush()

    print(f"Trial {trial_number}: Beep played & LED {current_led} turned on after 500 ms")

def on_start_button():
    asyncio.run_coroutine_threadsafe(start_recording_and_trial(), loop)

def on_stop_trial_button():
    asyncio.run_coroutine_threadsafe(stop_qtm_recording(), loop)

def on_end_button():
    export_to_excel()
    loop.call_soon_threadsafe(loop.stop)
    root.destroy()

# ------------------ EXCEL EXPORT ------------------
def export_to_excel():
    if not event_log:
        messagebox.showwarning("No Data", "No events to export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Trials"

    ws['A1'] = 'Trial'
    ws['B1'] = 'LED Lit'
    ws['C1'] = 'Timestamp'
    ws['D1'] = 'Event'
    ws['E1'] = 'Duration (ms)'

    for idx, (trial, led, timestamp, event, duration) in enumerate(event_log, start=2):
        ws[f"A{idx}"] = trial
        ws[f"B{idx}"] = led
        ws[f"C{idx}"] = timestamp
        ws[f"D{idx}"] = event
        if duration is not None:
            ws[f"E{idx}"] = duration

    filename = f"trial_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    messagebox.showinfo("Export Successful", f"Saved as {filename}")

# ------------------ GUI ------------------
def build_gui():
    global root, num_leds, led_pool

    root = tk.Tk()
    root.title("QTM + Arduino Trial Controller")

    # ask number of LEDs at startup
    num_leds = simpledialog.askinteger("Setup", "How many buttons/LEDs are connected? (1-4)", minvalue=1, maxvalue=4)
    led_pool = list(range(1, num_leds + 1))

    start_btn = tk.Button(root, text="Start Trial", command=on_start_button, width=25, height=3)
    start_btn.pack(pady=10, padx=20)

    stop_btn = tk.Button(root, text="Stop Trial", command=on_stop_trial_button, width=25, height=3)
    stop_btn.pack(pady=10, padx=20)

    end_btn = tk.Button(root, text="End & Save", command=on_end_button, width=25, height=3)
    end_btn.pack(pady=10, padx=20)

    root.protocol("WM_DELETE_WINDOW", on_end_button)
    root.mainloop()

# ------------------ MAIN ------------------
if __name__ == "__main__":
    build_gui()
