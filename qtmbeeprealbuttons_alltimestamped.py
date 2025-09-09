"""
To run this code, QTM (Qualisys Track Manager) must be set up and running on the same PC 
as the Arduino. The Arduino board must also be connected to this computer. 

First, verify the Arduino code is working, then upload it to the Elegoo Mega2560 R3, 
which should have up to 4 large buttons connected via a STEMMA QT I2C breakout. 

Make sure QTM is open and ready to record. When you run this script, you will be prompted 
to enter how many buttons are connected. The GUI will then appear with three buttons: 
Start Trial, Stop Trial, and End. 

Click 'Start Trial' to begin a trial. QTM will start recording, a beep sound will play, 
and a random button will light up. When the lit button is pressed, the light will turn off. 
Click 'Stop Trial' (do not click 'End') to stop recording the trial and save results. 

Each button will only be lit once per session. Continue until all trials are completed. 
When finished, click 'End' to close the GUI and export the results. The Excel file 
will contain the times when each button was pressed and released, along with the duration 
each button was held.
"""



# --- Integrated QTM + Arduino (Seesaw Buttons) + Excel Logger ---
# This program controls trials using an Arduino (with button inputs + LEDs),
# records data from QTM motion capture, plays beeps, and saves everything to Excel.

# Import required libraries
import asyncio          # For running asynchronous tasks (QTM + trial control)
import threading        # To run background tasks (like serial reader + event loop)
import serial           # For communicating with Arduino over USB/serial
import time             # For delays
import random           # For choosing random buttons
import numpy as np      # For handling audio data arrays
import sounddevice as sd # For playing sounds
import tkinter as tk    # For building a simple graphical user interface
from tkinter import simpledialog, messagebox  # GUI popups for input and alerts
from openpyxl import Workbook  # For saving results to Excel
from datetime import datetime  # For timestamps
import qtm             # For connecting to QTM motion capture system
import os              # For working with files and directories
import soundfile as sf  # For reading audio files
import subprocess
import pytz
from datetime import datetime


# ------------------ CONFIG ------------------
SERIAL_PORT = 'COM10'    # Port where Arduino is connected (change if needed)
BAUD_RATE = 115200       # Speed of serial communication (must match Arduino sketch)
WORK_DIR = r"C:\Users\AoMV Lab\ricki projects"  # Folder to save Excel logs
os.chdir(WORK_DIR)       # Change working directory to the above path

def sync_windows_time():
    """Force Windows to sync clock with time server."""
    try:
        output = subprocess.check_output(['w32tm', '/resync'], 
                                         shell=True, 
                                         stderr=subprocess.STDOUT, 
                                         text=True)
        print(f"Time sync success:\n{output}")
    except subprocess.CalledProcessError as e:
        print(f"Time sync failed:\n{e.output}")

CENTRAL_TZ = pytz.timezone('America/Chicago')

def now_central():
    """Get current time in Central Time."""
    return datetime.now(CENTRAL_TZ)


# ------------------ GLOBALS ------------------
qtm_connection = None                # Holds the QTM connection object
loop = asyncio.new_event_loop()      # Create a new asynchronous event loop

# Connect to Arduino
arduino = serial.Serial(port=SERIAL_PORT, baudrate=BAUD_RATE, timeout=1)
time.sleep(2)  # Wait a moment for Arduino to reset after connecting

# Data storage
event_log = []     # Stores logs of all events: (trial, button, timestamp, event, duration)
press_times = {}   # Keeps track of button press times to calculate duration
trial_number = 0   # Counter for trials
button_pool = []   # Buttons available to be chosen at random
current_button = None # Button currently active (lit up)
num_buttons = 0    # How many buttons are connected (user sets this at start)

# ------------------ EVENT LOOP ------------------
def start_event_loop():
    """Start the asyncio event loop so async tasks can run in background."""
    asyncio.set_event_loop(loop)
    loop.run_forever()

# Run event loop in a background thread
threading.Thread(target=start_event_loop, daemon=True).start()

# ------------------ QTM CONTROL ------------------
async def start_qtm_recording():
    """Connect to QTM and start recording."""
    global qtm_connection
    try:
        qtm_connection = await qtm.connect("127.0.0.1")  # Connect to QTM on local machine
        print("Connected to QTM.")
        await qtm_connection.take_control("")  # Take control of QTM
        await qtm_connection.start()           # Start recording
        print("Recording started.")
    except Exception as e:
        print(f"Failed to start QTM recording: {e}")
        qtm_connection = None

async def stop_qtm_recording():
    """Stop QTM recording and disconnect."""
    global qtm_connection
    if qtm_connection:
        try:
            await qtm_connection.stop()   # Stop recording
            print("Recording stopped.")
            qtm_connection.disconnect()   # Disconnect
            print("Disconnected from QTM.")
            qtm_connection = None
        except Exception as e:
            print(f"Failed to stop QTM recording: {e}")
            # If stopping fails, show an error popup
            def show_error():
                messagebox.showerror("Error", "Failed to stop recording.")
            root.after(0, show_error)

# ------------------ BEEP ------------------
def play_beep_blocking():
    """Play a beep sound (blocking until finished)."""
    filename = r"C:\Users\AoMV Lab\ricki projects\Silenceplus500hz1000mstone.wav"
    data, fs = sf.read(filename, dtype='float32')
    
    # Log beep actually starting
    event_log.append((trial_number, current_button, now_central().strftime('%H:%M:%S.%f')[:-3],
                      "Beep Started", None))
    
    sd.play(data, fs)
    sd.wait()


# ------------------ SERIAL READER ------------------
def read_serial():
    """Continuously read button press/release events from Arduino."""
    global trial_number, current_button
    while True:
        if arduino.in_waiting > 0:  # If there’s data waiting
            line = arduino.readline().decode(errors='ignore').strip()  # Read one line
            if not line:
                continue

            parts = line.split()  # Expect "event time"
            if len(parts) != 2:
                continue

            event = parts[0]        # e.g. "btn1_pressed"
            arduino_ms = int(parts[1])  # Time from Arduino in ms
            #system_time = datetime.now().strftime('%H:%M:%S.%f')[:-3]  # Current computer time
            system_time = now_central().strftime('%H:%M:%S.%f')[:-3]  # Central Time


            # Handle button press
            if "_pressed" in event:
                button = event.split("_")[1]  # Extract button number
                press_times[button] = arduino_ms  # Save press time
                event_text = f"#{button} - pressed"
                event_log.append((trial_number, current_button, system_time, event_text, None))
                print(f"[Trial {trial_number}] Button {current_button} [{system_time}] {event_text}")

                # Turn off LED once button is pressed
                if current_button is not None:
                    # Log LED off command
                    event_log.append((trial_number, current_button, now_central().strftime('%H:%M:%S.%f')[:-3],
                                      f"LED_{current_button}_OFF Command Sent", None))
                    arduino.write(f"LED_{current_button}_OFF\n".encode())
                    # Log LED actually turned off (approximation)
                    event_log.append((trial_number, current_button, now_central().strftime('%H:%M:%S.%f')[:-3],
                                          f"LED_{current_button}_Turned Off", None))

            # Handle button release
            elif "_released" in event:
                button = event.split("_")[1]
                if button in press_times:
                    duration = arduino_ms - press_times[button]  # How long button was held
                    event_text = f"#{button} - released"
                    event_log.append((trial_number, current_button, system_time, event_text, duration))
                    print(f"[Trial {trial_number}] Button {current_button} [{system_time}] {event_text} (Duration: {duration} ms)")
                    del press_times[button]  # Remove since released

# Start serial reader in background thread
serial_thread = threading.Thread(target=read_serial, daemon=True)
serial_thread.start()

# ------------------ TRIAL CONTROL ------------------
async def start_recording_and_trial():
    """Start a trial: record QTM, play beep, and light up random button."""
    global trial_number, current_button, button_pool

    if not button_pool:  # No buttons left to use
        messagebox.showinfo("Done", "All buttons have been used.")
        return

    trial_number += 1
    
    # Log QTM start command
    event_log.append((trial_number, None, now_central().strftime('%H:%M:%S.%f')[:-3],
                      "QTM Start Command Sent", None)) 
    
    await start_qtm_recording()  # Start QTM
    if qtm_connection is None:   # If connection failed, show error
        def show_error():
            messagebox.showerror("Error", "Failed to start recording.")
        root.after(0, show_error)
        return
        
    # Log QTM actually started
    event_log.append((trial_number, None, now_central().strftime('%H:%M:%S.%f')[:-3],
                      "QTM Recording Started", None))

    await asyncio.sleep(2)  # Small delay before beep

    # Choose a random button and remove it from pool
    current_button = random.choice(button_pool)
    button_pool.remove(current_button)


    # Before starting beep thread
    event_log.append((trial_number, current_button, now_central().strftime('%H:%M:%S.%f')[:-3],
                  "Beep Command Sent", None))
 
    # Play beep in separate thread so it doesn’t block
    beep_thread = threading.Thread(target=play_beep_blocking, daemon=True)
    beep_thread.start()

    # Wait before lighting LED
    await asyncio.sleep(1.0)
    
    # Command to turn on LED
    event_log.append((trial_number, current_button, now_central().strftime('%H:%M:%S.%f')[:-3],
                      f"LED_{current_button}_ON Command Sent", None))
    
    command = f"LED_{current_button}_ON\n"
    arduino.write(command.encode())
    arduino.flush()
    
    # Log LED actually lit (approximation if no Arduino confirmation)
    event_log.append((trial_number, current_button, now_central().strftime('%H:%M:%S.%f')[:-3],
                      f"LED_{current_button}_Lit", None))

    print(f"Trial {trial_number}: Beep played & Button {current_button} lit")

# GUI button actions
def on_start_button():
    asyncio.run_coroutine_threadsafe(start_recording_and_trial(), loop)

def on_stop_trial_button():
    asyncio.run_coroutine_threadsafe(stop_qtm_recording(), loop)

def on_end_button():
    export_to_excel()              # Save results
    loop.call_soon_threadsafe(loop.stop)  # Stop async loop
    root.destroy()                 # Close GUI

# ------------------ EXCEL EXPORT ------------------
def export_to_excel():
    """Save all logged events to Excel file."""
    if not event_log:
        messagebox.showwarning("No Data", "No events to export.")
        return

    wb = Workbook()       # Create new workbook
    ws = wb.active
    ws.title = "Trials"   # Rename sheet

    # Column headers
    ws['A1'] = 'Trial'
    ws['B1'] = 'Button Lit'
    ws['C1'] = 'Timestamp'
    ws['D1'] = 'Event'
    ws['E1'] = 'Duration (ms)'

    # Write each event to Excel
    for idx, (trial, button, timestamp, event, duration) in enumerate(event_log, start=2):
        ws[f"A{idx}"] = trial
        ws[f"B{idx}"] = button
        ws[f"C{idx}"] = timestamp
        ws[f"D{idx}"] = event
        if duration is not None:
            ws[f"E{idx}"] = duration

    # Save with timestamped filename
    filename = f"trial_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    messagebox.showinfo("Export Successful", f"Saved as {filename}")

# ------------------ GUI ------------------
def build_gui():
    """Build the graphical interface for controlling trials."""
    global root, num_buttons, button_pool

    root = tk.Tk()  # Create window
    root.title("QTM + Seesaw Trial Controller")

    # Ask user how many buttons are connected (1–4)
    num_buttons = simpledialog.askinteger("Setup", "How many Seesaw buttons are connected? (1-4)", minvalue=1, maxvalue=4)
    button_pool = list(range(1, num_buttons + 1))  # Initialize button pool

    # Add buttons to GUI
    start_btn = tk.Button(root, text="Start Trial", command=on_start_button, width=25, height=3)
    start_btn.pack(pady=10, padx=20)

    stop_btn = tk.Button(root, text="Stop Trial", command=on_stop_trial_button, width=25, height=3)
    stop_btn.pack(pady=10, padx=20)

    end_btn = tk.Button(root, text="End & Save", command=on_end_button, width=25, height=3)
    end_btn.pack(pady=10, padx=20)

    # Handle window close (same as pressing "End & Save")
    root.protocol("WM_DELETE_WINDOW", on_end_button)
    root.mainloop()

# ------------------ MAIN ------------------
if __name__ == "__main__":
    sync_windows_time()
    build_gui()  # Start program by opening GUI
